import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Comparador de Códigos", page_icon="🔍", layout="centered")

st.title("🔍 Comparador de Códigos")
st.markdown(
    "Subí el **Archivo A** (Excel con catálogo/stock) y el **Archivo B** (PDF con pedido) "
    "para encontrar los códigos que están en B pero **no** están en A."
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize(code: str) -> str:
    """Elimina todos los espacios del código para comparar sin importar el formato."""
    return code.replace(" ", "").upper()


def find_code_column(df):
    for col in df.columns:
        if str(col).strip().lower() in ("código","codigo","code","cod","códigos","codigos"):
            return df[col].dropna().astype(str).str.strip().str.upper(), str(col)
    return df.iloc[:, 0].dropna().astype(str).str.strip().str.upper(), str(df.columns[0])


def extract_codes_excel(file):
    df = pd.read_excel(file, dtype=str)
    series, col_name = find_code_column(df)
    # Normalizar: quitar espacios para comparación
    normalized = set(series.map(normalize))
    return normalized, col_name


def extract_codes_pdf(file):
    codes, seen = [], set()
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            words = page.extract_words()
            rows = {}
            for w in words:
                key = round(w["top"] / 3) * 3
                rows.setdefault(key, []).append(w)
            for top_key in sorted(rows):
                row_words = sorted(rows[top_key], key=lambda w: w["x0"])
                left_words = [w["text"] for w in row_words if w["x0"] < 115]
                if not left_words:
                    continue
                raw_code = " ".join(left_words).strip().upper()
                if not any(c.isdigit() for c in raw_code):
                    continue
                if ":" in raw_code or "--" in raw_code:
                    continue
                # Guardar el código original (con espacios) para mostrar,
                # y normalizado para comparar
                norm = normalize(raw_code)
                if norm not in seen:
                    seen.add(norm)
                    codes.append(raw_code)   # conserva formato original del PDF
    return codes


def build_excel_result(missing):
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos Faltantes"
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    cell_font   = Font(size=11, name="Arial")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill("solid", start_color="D6E4F0")
    ws.column_dimensions["A"].width = 30
    ws.row_dimensions[1].height = 22
    h = ws.cell(row=1, column=1, value="Código")
    h.fill, h.font, h.alignment, h.border = header_fill, header_font, center, border
    for i, code in enumerate(missing, start=2):
        c = ws.cell(row=i, column=1, value=code)
        c.font, c.alignment, c.border = cell_font, center, border
        if i % 2 == 0:
            c.fill = alt_fill
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────────

col1, col2 = st.columns(2)
with col1:
    st.subheader("📊 Archivo A — Excel")
    file_a = st.file_uploader("Catálogo / stock", type=["xlsx","xls"], key="file_a")
with col2:
    st.subheader("📄 Archivo B — PDF")
    file_b = st.file_uploader("Pedido / lista de compra", type=["pdf"], key="file_b")

st.divider()

if file_a and file_b:
    with st.spinner("Procesando archivos…"):
        try:
            codes_a, col_name = extract_codes_excel(file_a)
            codes_b = extract_codes_pdf(file_b)
        except Exception as e:
            st.error(f"❌ Error al leer los archivos: {e}")
            st.stop()

    missing = [c for c in codes_b if normalize(c) not in codes_a]
    found   = [c for c in codes_b if normalize(c) in codes_a]

    st.caption(f"ℹ️ Columna de códigos detectada en Excel: **{col_name}**")

    m1, m2, m3 = st.columns(3)
    m1.metric("Códigos en A (Excel)", f"{len(codes_a):,}")
    m2.metric("Códigos en B (PDF)",   f"{len(codes_b):,}")
    m3.metric("Faltan en A ❌", f"{len(missing):,}",
              delta=f"-{len(missing)}" if missing else "0",
              delta_color="inverse" if missing else "off")

    st.divider()

    if missing:
        st.error(f"### ❌ {len(missing)} código(s) del PDF **no están** en el Excel")
        st.dataframe(pd.DataFrame({"Código": missing}), use_container_width=True, hide_index=True)
        st.download_button(
            "⬇️ Descargar resultado (.xlsx)",
            data=build_excel_result(missing),
            file_name="codigos_faltantes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.success("✅ ¡Todos los códigos del PDF ya están en el Excel!")

    with st.expander(f"Ver detalle de los {len(codes_b)} código(s) del PDF"):
        rows = [{"Código": c, "Estado": "✅ Encontrado" if normalize(c) in codes_a else "❌ Faltante"} for c in codes_b]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

else:
    st.info("⬆️ Subí ambos archivos para comenzar la comparación.")

st.divider()
st.caption(
    "💡 **Tip:** El programa detecta automáticamente la columna de códigos en el Excel "
    "(busca columnas llamadas 'Código', 'Codigo', 'Code' o 'Cod'). "
    "En el PDF, extrae los códigos de la columna izquierda del pedido."
)
